#!/usr/bin/env python3
"""
Arrow Calendar Bot - Asistente de agenda personal en Telegram
"""

import os
import sqlite3
import logging
import json
import re
import tempfile
from datetime import datetime
from telegram import Update
from telegram.ext import (
    Application, CommandHandler, MessageHandler,
    filters, ContextTypes
)
import google.generativeai as genai
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)

TELEGRAM_TOKEN = os.environ.get("TELEGRAM_TOKEN", "")
GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY", "")

genai.configure(api_key=GEMINI_API_KEY)
gemini_model = genai.GenerativeModel('gemini-1.5-flash')

DB_PATH = "arrow_calendar.db"

# ==================== DATABASE ====================

def init_db():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('''
        CREATE TABLE IF NOT EXISTS reminders (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            description TEXT NOT NULL,
            reminder_time TEXT NOT NULL,
            created_at TEXT DEFAULT (datetime('now')),
            sent INTEGER DEFAULT 0
        )
    ''')
    conn.commit()
    conn.close()

def save_reminder(user_id: int, description: str, reminder_time: str):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute(
        'INSERT INTO reminders (user_id, description, reminder_time) VALUES (?, ?, ?)',
        (user_id, description, reminder_time)
    )
    conn.commit()
    conn.close()

def get_upcoming_reminders(user_id: int):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('''
        SELECT id, description, reminder_time FROM reminders
        WHERE user_id = ? AND sent = 0 AND datetime(reminder_time) >= datetime('now')
        ORDER BY reminder_time ASC LIMIT 20
    ''', (user_id,))
    rows = c.fetchall()
    conn.close()
    return rows

def get_all_reminders_for_excel(user_id: int):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('''
        SELECT id, description, reminder_time, created_at, sent FROM reminders
        WHERE user_id = ? ORDER BY reminder_time ASC
    ''', (user_id,))
    rows = c.fetchall()
    conn.close()
    return rows

def get_due_reminders():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('''
        SELECT id, user_id, description FROM reminders
        WHERE sent = 0 AND datetime(reminder_time) <= datetime('now')
    ''')
    rows = c.fetchall()
    conn.close()
    return rows

def mark_reminder_sent(reminder_id: int):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('UPDATE reminders SET sent = 1 WHERE id = ?', (reminder_id,))
    conn.commit()
    conn.close()

# ==================== AI ====================

def parse_with_gemini(text: str, current_dt: str) -> dict:
    prompt = f"""Eres un asistente de agenda personal. La fecha y hora actual es: {current_dt}

El usuario te envió: "{text}"

Responde SOLO con un JSON válido (sin markdown), con esta estructura exacta:
{{
  "intent": "create_reminder" | "view_agenda" | "export_excel" | "help" | "unknown",
  "description": "descripción del recordatorio o null",
  "reminder_datetime": "YYYY-MM-DD HH:MM:SS o null",
  "response_message": "mensaje amigable en español con emojis"
}}

Reglas:
- "agéndame", "recuérdame", "programa", "anota" → create_reminder
- "qué tengo", "mi agenda", "ver citas", "recordatorios" → view_agenda
- "excel", "exportar", "descargar lista" → export_excel
- "ayuda", "qué puedes hacer" → help
- Interpreta fechas relativas desde {current_dt}
- Si no dan hora, usa 09:00:00
- response_message debe confirmar la acción con emojis

Solo el JSON, nada más."""

    try:
        response = gemini_model.generate_content(prompt)
        raw = response.text.strip()
        raw = re.sub(r'```json\n?', '', raw)
        raw = re.sub(r'```\n?', '', raw)
        return json.loads(raw)
    except Exception as e:
        logger.error(f"Gemini parse error: {e}")
        return {
            "intent": "unknown",
            "response_message": "No entendí bien 😅 Puedes decirme:\n• _'Agéndame reunión mañana a las 3pm'_\n• _'Qué tengo esta semana'_\n• _'Mándame el excel'_"
        }

def transcribe_voice(audio_path: str) -> str:
    try:
        with open(audio_path, 'rb') as f:
            audio_data = f.read()
        response = gemini_model.generate_content([
            "Transcribe este audio en español. Devuelve solo el texto, sin explicaciones.",
            {"mime_type": "audio/ogg", "data": audio_data}
        ])
        return response.text.strip()
    except Exception as e:
        logger.error(f"Voice transcription error: {e}")
        return ""

# ==================== EXCEL ====================

def generate_excel(reminders: list, user_name: str) -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Arrow Calendar"

    header_fill = PatternFill(start_color="1A73E8", end_color="1A73E8", fill_type="solid")
    header_font = Font(name='Calibri', bold=True, color="FFFFFF", size=12)
    center = Alignment(horizontal='center', vertical='center')

    # Title
    ws.merge_cells('A1:E1')
    ws['A1'] = f"📅 Arrow Calendar — {user_name}"
    ws['A1'].font = Font(name='Calibri', bold=True, size=14, color="1A73E8")
    ws['A1'].alignment = center
    ws.row_dimensions[1].height = 30

    # Subtitle
    ws.merge_cells('A2:E2')
    ws['A2'] = f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A2'].font = Font(name='Calibri', italic=True, color="666666", size=10)
    ws['A2'].alignment = center

    # Headers
    for col, h in enumerate(['#', 'Descripción', 'Fecha y Hora', 'Estado', 'Creado el'], 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
    ws.row_dimensions[3].height = 25

    sent_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    pending_fill = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")

    for i, (rid, desc, rtime, created, sent) in enumerate(reminders, 1):
        row = i + 3
        fill = sent_fill if sent else pending_fill
        try:
            dt = datetime.strptime(rtime, '%Y-%m-%d %H:%M:%S')
            fmt_time = dt.strftime('%d/%m/%Y %H:%M')
        except:
            fmt_time = rtime
        try:
            ct = datetime.strptime(created, '%Y-%m-%d %H:%M:%S')
            fmt_created = ct.strftime('%d/%m/%Y %H:%M')
        except:
            fmt_created = created

        status = "✅ Enviado" if sent else "⏳ Pendiente"
        for col, val in enumerate([i, desc, fmt_time, status, fmt_created], 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = fill
            cell.font = Font(name='Calibri', size=11)
            cell.alignment = Alignment(vertical='center', wrap_text=True)
        ws.row_dimensions[row].height = 22

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 20

    filepath = f"/tmp/arrow_calendar_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(filepath)
    return filepath

# ==================== FORMAT AGENDA ====================

def format_agenda(reminders: list) -> str:
    if not reminders:
        return (
            "📭 No tienes recordatorios pendientes.\n\n"
            "Prueba escribirme:\n"
            "• _'Agéndame reunión mañana a las 10am'_\n"
            "• _'Recuérdame el dentista el viernes'_"
        )
    by_date = {}
    for rid, desc, rtime in reminders:
        try:
            dt = datetime.strptime(rtime, '%Y-%m-%d %H:%M:%S')
            date_key = dt.strftime('%A %d de %B').capitalize()
            time_str = dt.strftime('%H:%M')
        except:
            date_key = "Sin fecha"
            time_str = "?"
        by_date.setdefault(date_key, []).append((time_str, desc))

    text = "📅 *Tu Agenda — Arrow Calendar*\n"
    text += "─────────────────────\n\n"
    for date, events in by_date.items():
        text += f"📆 *{date}*\n"
        for time_str, desc in events:
            text += f"  🕐 {time_str} — {desc}\n"
        text += "\n"
    text += "─────────────────────\n"
    text += f"_Total: {len(reminders)} recordatorio(s) pendiente(s)_"
    return text

# ==================== HANDLERS ====================

async def start_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    await update.message.reply_text(
        f"¡Hola {user.first_name}! 👋\n\n"
        "Soy *Arrow Calendar* 📅, tu asistente personal de agenda.\n\n"
        "*¿Qué puedo hacer?*\n"
        "📌 Agendar recordatorios con lenguaje natural\n"
        "🎤 Entender mensajes de voz\n"
        "📊 Exportar tu agenda a Excel\n"
        "⏰ Enviarte recordatorios automáticos\n\n"
        "*Ejemplos:*\n"
        "• _'Agéndame dentista el viernes a las 3pm'_\n"
        "• _'Recuérdame llamar a mamá mañana a las 6'_\n"
        "• _'Qué tengo esta semana'_\n"
        "• _'Mándame el excel'_\n\n"
        "¡Escríbeme lo que necesites! 🚀",
        parse_mode='Markdown'
    )

async def agenda_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    reminders = get_upcoming_reminders(update.effective_user.id)
    await update.message.reply_text(format_agenda(reminders), parse_mode='Markdown')

async def excel_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    await update.message.reply_text("⏳ Generando tu Excel, un momento...")
    reminders = get_all_reminders_for_excel(user.id)
    if not reminders:
        await update.message.reply_text("📭 No tienes recordatorios guardados todavía.")
        return
    filepath = generate_excel(reminders, user.first_name or "Usuario")
    with open(filepath, 'rb') as f:
        await update.message.reply_document(
            document=f,
            filename=f"Arrow_Calendar_{datetime.now().strftime('%d%m%Y')}.xlsx",
            caption=f"📊 Tu agenda completa — {len(reminders)} recordatorio(s)\n_Arrow Calendar_ ✅",
            parse_mode='Markdown'
        )

async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    text = update.message.text
    current_dt = datetime.now().strftime('%Y-%m-%d %H:%M:%S (%A %d de %B de %Y)')
    parsed = parse_with_gemini(text, current_dt)
    intent = parsed.get('intent', 'unknown')

    if intent == 'create_reminder':
        description = parsed.get('description') or text
        reminder_time = parsed.get('reminder_datetime')
        if reminder_time:
            save_reminder(user.id, description, reminder_time)
            try:
                dt = datetime.strptime(reminder_time, '%Y-%m-%d %H:%M:%S')
                formatted = dt.strftime('%A %d de %B a las %H:%M').capitalize()
            except:
                formatted = reminder_time
            await update.message.reply_text(
                f"✅ *¡Recordatorio guardado!*\n\n"
                f"📌 {description}\n"
                f"⏰ {formatted}\n\n"
                f"_Te avisaré a tiempo_ 🔔",
                parse_mode='Markdown'
            )
        else:
            await update.message.reply_text(
                parsed.get('response_message', '¿Puedes especificar la fecha y hora? 😊'),
                parse_mode='Markdown'
            )

    elif intent == 'view_agenda':
        reminders = get_upcoming_reminders(user.id)
        await update.message.reply_text(format_agenda(reminders), parse_mode='Markdown')

    elif intent == 'export_excel':
        await excel_command(update, context)

    elif intent == 'help':
        await update.message.reply_text(
            "📖 *Arrow Calendar — Ayuda*\n\n"
            "🗣️ Escríbeme en lenguaje natural:\n\n"
            "➕ *Crear recordatorio:*\n"
            "• 'Agéndame reunión mañana a las 10am'\n"
            "• 'Recuérdame el dentista el viernes 3pm'\n\n"
            "📅 *Ver agenda:*\n"
            "• 'Qué tengo esta semana'\n"
            "• 'Mis recordatorios'\n"
            "• /agenda\n\n"
            "📊 *Exportar Excel:*\n"
            "• 'Mándame el excel'\n"
            "• /excel\n\n"
            "🎤 *Voz:*\n"
            "Envíame un audio y lo entiendo automáticamente",
            parse_mode='Markdown'
        )

    else:
        await update.message.reply_text(
            parsed.get('response_message',
                '¿En qué te ayudo? 😊\nEscribe /help para ver todo lo que puedo hacer.'),
            parse_mode='Markdown'
        )

async def handle_voice(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("🎤 Escuchando tu audio...")
    voice = update.message.voice
    file = await context.bot.get_file(voice.file_id)
    with tempfile.NamedTemporaryFile(suffix='.ogg', delete=False) as tmp:
        await file.download_to_drive(tmp.name)
        transcribed = transcribe_voice(tmp.name)
    if transcribed:
        await update.message.reply_text(
            f"🎤 _Entendí:_ \"{transcribed}\"",
            parse_mode='Markdown'
        )
        update.message.text = transcribed
        await handle_text(update, context)
    else:
        await update.message.reply_text("❌ No pude entender el audio. ¿Puedes escribirlo?")

# ==================== REMINDER CHECKER JOB ====================

async def check_reminders(context: ContextTypes.DEFAULT_TYPE):
    due = get_due_reminders()
    for reminder_id, user_id, description in due:
        try:
            await context.bot.send_message(
                chat_id=user_id,
                text=f"🔔 *¡Recordatorio Arrow Calendar!*\n\n📌 {description}\n\n_¡No lo olvides!_ ✅",
                parse_mode='Markdown'
            )
            mark_reminder_sent(reminder_id)
            logger.info(f"Sent reminder {reminder_id} to user {user_id}")
        except Exception as e:
            logger.error(f"Error sending reminder {reminder_id}: {e}")

# ==================== MAIN ====================

def main():
    init_db()
    app = Application.builder().token(TELEGRAM_TOKEN).build()

    app.add_handler(CommandHandler("start", start_command))
    app.add_handler(CommandHandler("help", lambda u, c: handle_text(u, c)))
    app.add_handler(CommandHandler("agenda", agenda_command))
    app.add_handler(CommandHandler("excel", excel_command))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text))
    app.add_handler(MessageHandler(filters.VOICE, handle_voice))

    # Check reminders every 60 seconds
    app.job_queue.run_repeating(check_reminders, interval=60, first=10)

    logger.info("🚀 Arrow Calendar Bot is running!")
    app.run_polling(allowed_updates=Update.ALL_TYPES)

if __name__ == '__main__':
    main()
